import os, hashlib
from openpyxl import Workbook, load_workbook

DB_FILE = "users.xlsx"

def _hash(pw: str) -> str:
    return hashlib.sha256(pw.encode("utf-8")).hexdigest()

def init_db():
    if not os.path.exists(DB_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        ws.append(["name", "age", "password_hash"])
        wb.save(DB_FILE)

def register_user(name: str, age: str, password: str):
    init_db()
    wb = load_workbook(DB_FILE)
    ws = wb["Users"]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] == name:
            return False, "Name already exists."
    ws.append([name, age, _hash(password)])
    wb.save(DB_FILE)
    return True, "Account created."

def find_by_password(password: str):
    """Returns list of dicts {name, age} that match the given password (per your spec)."""
    init_db()
    target = _hash(password)
    wb = load_workbook(DB_FILE)
    ws = wb["Users"]
    matches = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[2] == target:
            matches.append({"name": r[0], "age": r[1]})
    return matches
